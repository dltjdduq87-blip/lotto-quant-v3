"""One-time import of the user-supplied authoritative history workbook
(회차별 당첨번호, round 1 - N) into the SQLite store.

Expected sheet layout (Korean lotto stats export):
    col A: No (row index, descending: 1 = most recent round)
    col B: 회차 (round number)
    col C-H: 당첨번호 1-6 (six winning numbers, merged header)
    col I: 보너스 (bonus number)
    col J: 등수 (rank label, e.g. "1등") -- ignored
    col K: 당첨자수 -- ignored
    col L: 1인당 당첨금액 -- ignored
"""
from datetime import date, timedelta
from pathlib import Path

import openpyxl

from lotto_quant_v3.config.settings import FIRST_DRAW_DATE
from lotto_quant_v3.data.collector import DrawResult

_FIRST_DRAW = date.fromisoformat(FIRST_DRAW_DATE)


def round_to_date(round_no: int) -> str:
    """Lotto 6/45 has drawn every Saturday without exception since round 1
    (2002-12-07); verified against round 1236 -> 2026-08-08."""
    return (_FIRST_DRAW + timedelta(weeks=round_no - 1)).isoformat()


def import_excel(path: str | Path) -> list[DrawResult]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    results: list[DrawResult] = []
    for row in rows:
        if row[1] is None:
            continue
        round_no = int(row[1])
        numbers = tuple(int(row[i]) for i in range(2, 8))
        bonus = int(row[8])
        draw = DrawResult(round_no, draw_date=round_to_date(round_no), numbers=numbers, bonus=bonus)
        if not draw.is_structurally_valid():
            raise ValueError(f"invalid row for round {round_no}: {numbers} bonus={bonus}")
        results.append(draw)

    round_numbers = sorted(d.round for d in results)
    expected = list(range(round_numbers[0], round_numbers[-1] + 1))
    missing = sorted(set(expected) - set(round_numbers))
    if missing:
        raise ValueError(f"gap(s) in imported rounds, missing: {missing}")

    return sorted(results, key=lambda d: d.round)
